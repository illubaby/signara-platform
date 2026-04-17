#!/depot/Python/Python-3.8.0/bin/python
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook


INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1F]')


def sanitize_filename(name: str) -> str:
    """Return a filesystem-safe file name stem from a sheet name."""
    sanitized = INVALID_FILENAME_CHARS.sub("_", name).strip().strip(".")
    return sanitized or "sheet"


def unique_output_path(output_dir: Path, base_name: str) -> Path:
    """Return a unique CSV path, appending a numeric suffix when needed."""
    candidate = output_dir / f"{base_name}.csv"
    if not candidate.exists():
        return candidate

    suffix = 2
    while True:
        candidate = output_dir / f"{base_name}_{suffix}.csv"
        if not candidate.exists():
            return candidate
        suffix += 1


def export_sheets_to_csv(
    xlsx_path: Path,
    output_dir: Path,
    encoding: str = "utf-8-sig",
    keep_empty_rows: bool = False,
) -> list[Path]:
    """Export all workbook sheets to individual CSV files."""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Input file not found: {xlsx_path}")

    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    exported_files: list[Path] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        safe_name = sanitize_filename(sheet_name)
        csv_path = unique_output_path(output_dir, safe_name)

        with csv_path.open("w", newline="", encoding=encoding) as f:
            writer = csv.writer(f)
            pending_empty_rows = 0
            for row in worksheet.iter_rows(values_only=True):
                cleaned_row = ["" if value is None else value for value in row]

                # openpyxl may report an oversized used range due to formatting.
                # Remove trailing empty columns and do not write trailing empty rows.
                while cleaned_row and cleaned_row[-1] == "":
                    cleaned_row.pop()

                if not cleaned_row:
                    if keep_empty_rows:
                        pending_empty_rows += 1
                    continue

                if keep_empty_rows:
                    while pending_empty_rows > 0:
                        writer.writerow([])
                        pending_empty_rows -= 1

                writer.writerow(cleaned_row)

        exported_files.append(csv_path)

    workbook.close()
    return exported_files


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Export each sheet in an XLSX workbook to a separate CSV file.",
        add_help=True,
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python xlsx_sheets_to_csv.py input.xlsx out_dir\n"
            "  python xlsx_sheets_to_csv.py input.xlsx out_dir --encoding utf-8"
        ),
    )
    parser.add_argument("xlsx_file", type=Path, help="Path to the input .xlsx file")
    parser.add_argument("output_dir", type=Path, help="Directory to write CSV files into")
    parser.add_argument(
        "--encoding",
        default="utf-8-sig",
        help="CSV text encoding (default: utf-8-sig)",
    )
    parser.add_argument(
        "--keep-empty-rows",
        action="store_true",
        help="Preserve empty rows from the worksheet (default: skip empty rows)",
    )
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()

    exported = export_sheets_to_csv(
        args.xlsx_file,
        args.output_dir,
        args.encoding,
        keep_empty_rows=args.keep_empty_rows,
    )
    print(f"Exported {len(exported)} sheet(s):")
    for path in exported:
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())