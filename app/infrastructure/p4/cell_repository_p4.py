"""Perforce (P4) implementation of CellRepository.

Parses ProjectInfo.xlsx from Perforce depot to extract cell information.
"""
from __future__ import annotations

import subprocess
import tempfile
import os
import time
from typing import List, Dict, Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None

from app.domain.cell.entities import Cell
from app.domain.cell.repositories import CellRepository


class CellRepositoryP4(CellRepository):
    """Concrete P4 implementation for fetching cell data from ProjectInfo.xlsx."""
    
    def __init__(self) -> None:
        pass
    
    def list_cells(self, project: str, subproject: str, refresh: bool = False) -> List[Cell]:
        """
        Fetch and parse ProjectInfo.xlsx from Perforce depot.
        
        Args:
            project: The project name.
            subproject: The subproject name.
            refresh: This argument is ignored by this implementation but required by the protocol.

        Returns list of Cell entities populated from both worksheets:
        - List_cell_INT: All NT cells
        - List_cell_ETM: Mixed SIS/NT cells based on "Tool" column
        """
        depot_path = f"//wwcad/msip/projects/ucie/{project}/{subproject}/pcs/design/timing/ProjectInfo.xlsx"
        
        cells: List[Cell] = []
        
        try:
            # Fetch from Perforce
            result = subprocess.run(
                ["p4", "print", "-q", depot_path],
                capture_output=True,
                timeout=30
            )
            
            if result.returncode != 0 or not result.stdout:
                raise RuntimeError(f"p4 print failed for {depot_path}")
            
            if openpyxl is None:
                return []
            
            # Write to temp file
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(result.stdout)
                tmp_path = tmp.name
            
            try:
                wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                
                # Parse List_cell_INT sheet (all NT cells)
                if "List_cell_INT" in wb.sheetnames:
                    cells.extend(self._parse_int_sheet(wb["List_cell_INT"]))
                
                # Parse List_cell_ETM sheet (mixed SIS/NT cells)
                if "List_cell_ETM" in wb.sheetnames:
                    cells.extend(self._parse_etm_sheet(wb["List_cell_ETM"]))
                
                # Remove duplicates based on ckt_macros
                seen = set()
                unique_cells = []
                for cell in cells:
                    if cell.ckt_macros not in seen:
                        seen.add(cell.ckt_macros)
                        unique_cells.append(cell)
                cells = unique_cells
                
                wb.close()
            finally:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
        
        except Exception:
            # In case of any error, return an empty list.
            # The cached version (if any) will be used by the decorator.
            return []
        
        return cells
    
    def _parse_int_sheet(self, ws) -> List[Cell]:
        """Parse List_cell_INT worksheet (all NT cells)."""
        cells: List[Cell] = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            
            # Check if first column has data
            if row[0] is None or str(row[0]).strip() == "":
                continue
            
            cell_name = str(row[0]).strip()
            
            # Skip header-like entries (but not actual cell names starting with these)
            lower_name = cell_name.lower()
            if lower_name in ("cell", "name", "cell name", "ckt macros") or cell_name.startswith("#"):
                continue
            
            # Create Cell entity with SIS type
            cells.append(Cell(
                ckt_macros=cell_name,
                tool="nt",
            ))
        
        return cells
    
    def _parse_etm_sheet(self, ws) -> List[Cell]:
        """Parse List_cell_ETM worksheet (mixed SIS/NT cells based on Tool column)."""
        cells: List[Cell] = []
        
        # Find Tool column index
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        tool_idx = None
        
        if header:
            for i, h in enumerate(header):
                if h and str(h).strip().lower() in ("tool", "ckt macros"):
                    if str(h).strip().lower() == "tool":
                        tool_idx = i
                        break
        
        # Parse data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            
            # Check if first column has data
            if row[0] is None or str(row[0]).strip() == "":
                continue
            
            cell_name = str(row[0]).strip()
            
            # Skip header-like entries (but not actual cell names starting with these)
            lower_name = cell_name.lower()
            if lower_name in ("cell", "name", "cell name", "ckt macros") or cell_name.startswith("#"):
                continue
            
            # Determine cell type from Tool column
            tool_value = None
            if tool_idx is not None and len(row) > tool_idx and row[tool_idx]:
                tool_value = str(row[tool_idx]).strip()
                
                # Skip N/A entries
                if tool_value.lower() in ("na", "n/a", "n.a.", "n.a"):
                    continue
            
            # Default to NT if tool not specified or not SiS
            if not tool_value:
                tool_value = "nt"
            elif "sis" not in tool_value.lower():
                tool_value = "nt"
            else:
                tool_value = "sis"
            
            # Create Cell entity
            cells.append(Cell(
                ckt_macros=cell_name,
                tool=tool_value,
            ))
        
        return cells
    
    def _safe_str(self, row: tuple, idx: int) -> Optional[str]:
        """Safely extract string value from row at index."""
        if len(row) > idx and row[idx] is not None:
            val = str(row[idx]).strip()
            return val if val else None
        return None
    
    def _safe_int(self, row: tuple, idx: int) -> Optional[int]:
        """Safely extract integer value from row at index."""
        if len(row) > idx and row[idx] is not None:
            try:
                return int(row[idx])
            except (ValueError, TypeError):
                return None
        return None


__all__ = ["CellRepositoryP4"]
