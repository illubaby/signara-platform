"""Perforce (P4) implementation of SafPerforceRepository (Phase 3).

Encapsulates all subprocess calls to `p4` for SAF logic.
Parsing kept intentionally simple; future phases may harden error handling.
"""
from __future__ import annotations

import subprocess
import re
import tempfile
import os
import time
from typing import Dict, Optional, List

try:
    import openpyxl  # optional; warn gracefully if missing
except ImportError:  # pragma: no cover - test suite will mock when absent
    openpyxl = None

from app.domain.saf.repositories import SafPerforceRepository


class SafPerforceRepositoryP4(SafPerforceRepository):
    CACHE_INTERVAL = 300  # seconds

    def __init__(self) -> None:
        self._cells_cache: Dict[str, Dict] = {}

    # ---- helpers ----
    def get_last_change(self, depot_path: str) -> Optional[str]:
        try:
            result = subprocess.run([
                "p4", "changes", "-m1", "-s", "submitted", depot_path
            ], capture_output=True, text=True, timeout=10)
            if result.returncode == 0 and result.stdout:
                m = re.match(r'^Change (\d+)', result.stdout)
                if m:
                    return m.group(1)
        except Exception:
            return None
        return None

    def list_depot_cells(self, project: str, subproject: str) -> Dict[str, str]:
        depot_path = f"//wwcad/msip/projects/ucie/{project}/{subproject}/pcs/design/timing/ProjectInfo.xlsx"
        now = time.time()
        cache = self._cells_cache.get(depot_path)
        if cache:
            if (now - cache.get("last_check", 0)) < self.CACHE_INTERVAL:
                return cache["cells"]
        cells: Dict[str, str] = {}
        try:
            result = subprocess.run(["p4", "print", "-q", depot_path], capture_output=True, timeout=30)
            if result.returncode != 0 or not result.stdout:
                raise RuntimeError("p4 print failed")
            if openpyxl is None:
                if cache:
                    cache["last_check"] = now
                    return cache["cells"]
                return {}
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(result.stdout)
                tmp_path = tmp.name
            try:
                wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                if "List_cell_ETM" in wb.sheetnames:
                    ws = wb["List_cell_ETM"]
                    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    tool_idx = None
                    if header:
                        for i, h in enumerate(header):
                            if h and str(h).strip().lower() == "tool":
                                tool_idx = i; break
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            name = str(row[0]).strip()
                            if name and not name.lower().startswith(("cell", "name", "#")):
                                cell_type = None
                                if tool_idx is not None and len(row) > tool_idx and row[tool_idx]:
                                    tv = str(row[tool_idx]).strip().lower()
                                    if tv in ("na", "n/a", "n.a.", "n.a"):
                                        continue
                                    if "sis" in tv:
                                        cell_type = "sis"
                                    else:
                                        cell_type = "nt"
                                elif tool_idx is not None:
                                    continue
                                else:
                                    cell_type = "nt"
                                if cell_type:
                                    cells[name] = cell_type
                wb.close()
            finally:
                try: os.unlink(tmp_path)
                except Exception: pass
        except Exception:
            if cache:
                return cache["cells"]
            return {}
        last_change = self.get_last_change(depot_path)
        self._cells_cache[depot_path] = {"cells": cells, "last_change": last_change, "last_check": time.time()}
        return cells

    def list_int_depot_cells(self, wb) -> Dict[str, str]:
        """Extract cells from List_cell_INT worksheet marking them as SiS.

        Returns a dict mapping cell name -> "sis".
        Workbook object is expected to be an openpyxl workbook already loaded.
        """
        result: Dict[str, str] = {}
        try:
            if wb and "List_cell_INT" in wb.sheetnames:
                ws = wb["List_cell_INT"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        name = str(row[0]).strip()
                        if name and not name.lower().startswith(("cell", "name", "#")):
                            result[name] = "sis"
        except Exception:
            # Silently ignore parse errors; caller will proceed with what it has.
            return result
        return result

    def list_int_cells(self, project: str, subproject: str) -> List[str]:
        """Return INT cells (treated as SiS) from ProjectInfo.xlsx List_cell_INT sheet."""
        depot_path = f"//wwcad/msip/projects/ucie/{project}/{subproject}/pcs/design/timing/ProjectInfo.xlsx"
        try:
            result = subprocess.run(["p4", "print", "-q", depot_path], capture_output=True, timeout=30)
            if result.returncode != 0 or not result.stdout or openpyxl is None:
                return []
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(result.stdout)
                tmp_path = tmp.name
            try:
                wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                mapping = self.list_int_depot_cells(wb)
                wb.close()
                # Append '_int' suffix to each cell (avoid double suffix if already present)
                return sorted([name if name.endswith('_int') else f"{name}_int" for name in mapping.keys()])
            finally:
                try: os.unlink(tmp_path)
                except Exception: pass
        except Exception:
            return []

    def read_configure_pvt(self, project: str, subproject: str, internal: bool = False) -> List[str]:
        subdir = "internal/" if internal else ""
        depot_path = f"//wwcad/msip/projects/ucie/{project}/{subproject}/pcs/design/timing/sis/common_source/{subdir}configure_pvt.tcl"
        pvts: List[str] = []
        try:
            result = subprocess.run(["p4", "print", "-q", depot_path], capture_output=True, text=True, timeout=30)
            if result.returncode != 0:
                return []
            for line in result.stdout.splitlines():
                line = line.strip()
                if line.startswith("set_opc_process "):
                    parts = line.split()
                    if len(parts) >= 3 and parts[2] == '{':
                        pvts.append(parts[1])
        except subprocess.TimeoutExpired:
            return []
        except Exception:
            return []
        return pvts

    def sync_inst_file(self, project: str, subproject: str, cell: str, cell_type: str) -> Dict:
        if cell_type == "nt":
            file_name = "alphaNT.config"
            depot_path = f"//wwcad/msip/projects/ucie/{project}/{subproject}/design/timing/nt/{cell}/..."
        else:
            file_name = "*.inst"
            depot_path = f"//wwcad/msip/projects/ucie/{project}/{subproject}/design/timing/sis/{cell}/..."
        # We just run sync; caller inspects filesystem afterwards (higher layer). For now just capture output.
        try:
            result = subprocess.run(["p4", "sync", depot_path], capture_output=True, text=True, timeout=60)
            ok = result.returncode == 0
            return {
                "synced": ok,
                "file": file_name,
                "depot": depot_path,
                "stdout": (result.stdout or "")[:10000],
                "stderr": (result.stderr or "")[:8000],
            }
        except subprocess.TimeoutExpired:
            return {"synced": False, "file": file_name, "depot": depot_path, "error": "timeout"}
        except Exception as e:
            return {"synced": False, "file": file_name, "depot": depot_path, "error": str(e)}

__all__ = ["SafPerforceRepositoryP4"]
